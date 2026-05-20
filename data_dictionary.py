"""
Run this script once from Terminal to generate the Excel file:

    python3 data_dictionary.py

It will create "Mews_Dashboard_Data_Dictionary.xlsx" in your mews-dashboard folder.
You need openpyxl installed: pip3 install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    # (Graph Title, Metrics Shown, Catalog / Table, Calculations Applied, Comments)
    (
        "Key Metrics — Avg ADR",
        "Average Daily Rate in EUR",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(adr_eur) across all properties and days in the selected date range",
        "Pre-calculated in the mart as total adjusted net accommodation revenue / occupied rooms. No revenue outliers included (>€200k flagged).",
    ),
    (
        "Key Metrics — Avg Occupancy",
        "Occupancy % (rooms occupied / rooms available)",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(num_directly_occupied / num_available) * 100 across all properties and days",
        "Blocked rooms are excluded from available rooms. Occupancy inheritance from parent resources is not applied to avoid ADR distortion.",
    ),
    (
        "Key Metrics — Avg RevPAR",
        "Revenue per Available Room in EUR",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(revpar_eur) across all properties and days in the selected date range",
        "Pre-calculated in the mart as total adjusted net accommodation revenue / available rooms.",
    ),
    (
        "Key Metrics — Total Reservations",
        "Count of distinct reservations",
        "product.facts.fct_reservations + product.dimensions.dim_pms_properties",
        "COUNT(DISTINCT reservation_id) where reservation_state_code NOT IN (4) [i.e. not cancelled] and is_reservation_deleted = FALSE",
        "Filtered to subscribed/selected customer statuses. Date filter applied on reservation_planned_start_at.",
    ),
    (
        "Historical Performance — ADR (line chart)",
        "Daily average ADR in EUR per year, shown as overlapping lines",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(adr_eur) grouped by calendar date, then 7-day rolling average applied in Python per year. X-axis is day-of-year so multiple years overlay.",
        "Rolling average uses min_periods=1 so early days in the year use fewer than 7 data points. Each year is a separate coloured line.",
    ),
    (
        "Historical Performance — Occupancy (line chart)",
        "Daily average occupancy % per year",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(occupied/available)*100 grouped by calendar date, then 7-day rolling average per year in Python",
        "Same methodology as ADR line chart above.",
    ),
    (
        "Historical Performance — RevPAR (line chart)",
        "Daily average RevPAR in EUR per year",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property",
        "AVG(revpar_eur) grouped by calendar date, then 7-day rolling average per year in Python",
        "Same methodology as ADR line chart above.",
    ),
    (
        "On The Books — ADR (line chart)",
        "OTB ADR in EUR: This Year vs Last Year, for next 9 months",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property_on_the_books",
        "Two snapshot dates queried: most recent available (e.g. 2026-03-15) and same date one year prior (2025-03-15). X-axis is days_ahead = DATEDIFF(on_the_books_date, snapshot_date), range 0–274 (~9 months). AVG(adr_eur) grouped by days_ahead. 7-day rolling average applied in Python.",
        "This Year line = what is currently booked for each future date. Last Year line = what was booked for each equivalent future date as of the same snapshot one year ago. Allows like-for-like booking window comparison.",
    ),
    (
        "On The Books — Occupancy (line chart)",
        "OTB Occupancy %: This Year vs Last Year, for next 9 months",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property_on_the_books",
        "Same as OTB ADR but using AVG(occupied/available)*100",
        "Same methodology and caveats as OTB ADR.",
    ),
    (
        "On The Books — RevPAR (line chart)",
        "OTB RevPAR in EUR: This Year vs Last Year, for next 9 months",
        "product.marts.mrt_daily_resource_and_revenue_metrics_per_property_on_the_books",
        "Same as OTB ADR but using AVG(revpar_eur)",
        "Same methodology and caveats as OTB ADR.",
    ),
    (
        "Reservations by Lead Time (bar chart)",
        "% of reservations by number of days between booking creation and check-in date",
        "product.facts.fct_reservations + product.dimensions.dim_pms_properties",
        "DATEDIFF(reservation_planned_start_at, reservation_created_at) bucketed into 8 groups: 0 (same day), 1–3, 4–7, 8–14, 15–30, 31–60, 61–90, 90+ days. Each bucket shown as % of total reservations.",
        "Cancelled reservations excluded (state_code = 4). Deleted reservations excluded.",
    ),
    (
        "Reservations by Length of Stay (bar chart)",
        "% of reservations by number of nights",
        "product.facts.fct_reservations + product.dimensions.dim_pms_properties",
        "DATEDIFF(reservation_planned_end_at, reservation_planned_start_at) bucketed into 6 groups: 1, 2, 3, 4–7, 8–14, 15+ nights. Each bucket shown as % of total.",
        "Uses planned dates, not actual check-in/out dates. Cancelled and deleted reservations excluded.",
    ),
    (
        "Reservations by Group Size (bar chart)",
        "% of reservations by number of guests (person_count)",
        "product.facts.fct_reservations + product.dimensions.dim_pms_properties",
        "person_count bucketed into 5 groups: 1, 2, 3, 4–6, 7+ guests. Each bucket shown as % of total.",
        "Reservations where person_count is NULL are excluded. This is the count of guests on a single reservation, not a group booking count.",
    ),
    (
        "Reservations by Channel (pie chart)",
        "% of reservations by booking channel",
        "product.facts.fct_reservations + product.dimensions.dim_reservation_attributes + product.dimensions.dim_pms_properties",
        "reservation_origin and reservation_commander_origin used to bucket into 3 groups: Third-Party Channels (ChannelManager + Connector), Online Direct (Distributor + Navigator + Commander/Website), Offline Direct (all other Commander sub-origins). Import excluded entirely.",
        "Commander origin is a sub-classification used when origin = Commander. See dim_reservation_attributes for full list of values.",
    ),
    (
        "Payment Type Breakdown (pie chart)",
        "% of payment transactions by settlement type",
        "product.marts.mrt_billing_open_order_items_and_payments + product.dimensions.dim_pms_properties",
        "item_subtype filtered to item_type = 'Payment', then bucketed: CreditCardPayment + ExternalPayment + AlternativePayment → Card & Digital Payments; CashPayment → Cash; GhostPayment → Invoice & Corporate Billing; TaxDeductedPayment → Bank Transfer & Settlement. SUM(count_open_items_on_date) used as the volume metric.",
        "This table tracks open (unbilled) items. Card brand detail (Visa/Mastercard/Amex) is not available in this data layer — raw payment transaction table would be needed.",
    ),
    (
        "Card vs Other Payments (pie chart)",
        "% split between Card & Digital vs all other payment types",
        "product.marts.mrt_billing_open_order_items_and_payments",
        "Derived from Payment Type Breakdown: Card & Digital Payments bucket vs sum of all other buckets",
        "Simplified view of the payment type pie. No additional data pulled.",
    ),
    (
        "Pre-Stay OCI (pie chart)",
        "% of reservations where guest completed online check-in vs did not",
        "product.facts.fct_reservation_checkins__online + product.facts.fct_reservations + product.dimensions.dim_pms_properties",
        "COUNT(*) of completed online check-ins (online_check_in_finished_at is not null and within date range). OCI No = Total Reservations KPI minus OCI count.",
        "online_check_in_finished_at is used (not started_at) to ensure only completed check-ins are counted. A reservation can appear in both OCI and at-hotel check-in methods.",
    ),
    (
        "At Hotel Check-in Method (pie chart)",
        "% split between Kiosk, Digital Key, and Front Desk (PMS) check-ins",
        "fct_reservation_checkins__kiosk, fct_reservation_checkins__connector, mrt_digital_key_events_per_property_date, fct_front_desk_operations_checkins",
        "Four queries unioned and summed by method: Kiosk = kiosk table + connector table (third-party kiosk) summed together. Digital Key = SUM(count_keycards_activated). Front Desk = COUNT(*) from front_desk_operations_checkins.",
        "Connector check-ins are assumed to be third-party kiosk solutions and merged into the Kiosk bucket. A single reservation may appear in multiple methods if checked in via more than one channel.",
    ),
    (
        "Upsells by Channel (pie chart)",
        "% of upsells sold through each channel: Booking Engine, Kiosk, Online Check-in, Front Desk",
        "product.marts.mrt_daily_upsells_channel_shares_and_weights_per_accounting_category + product.dimensions.dim_pms_properties",
        "SUM of count_booking_engine_upsells, count_kiosk_upsells, count_guest_portal_upsells, count_front_desk_upsells across all properties and dates. Each channel shown as % of total.",
        "Note: a single reservation can have upsells across multiple channels, so totals may exceed total reservation count. guest_portal = Online Check-in (OCI) in Mews terminology.",
    ),
    (
        "Avg Upsell Value by Channel (bar chart)",
        "Average upsell revenue per reservation in EUR, by channel",
        "product.marts.mrt_daily_upsells_channel_shares_and_weights_per_accounting_category + product.dimensions.dim_pms_properties",
        "AVG of avg_booking_engine_upsell_value_per_reservation_eur, avg_kiosk_upsell_value_per_reservation_eur, avg_guest_portal_upsell_value_per_reservation_eur, avg_front_desk_upsell_value_per_reservation_eur across all properties and dates.",
        "Pre-calculated averages in the mart are averaged again across properties — this is an average of averages and may differ slightly from computing directly from raw transactions.",
    ),
    (
        "Top Upsell Categories — Volume (bar chart)",
        "% of upsells by accounting category classification (top 10)",
        "product.marts.mrt_daily_upsells_by_product + product.dimensions.dim_pms_properties",
        "SUM(count_total_upsells) grouped by accounting_category_classification. Payments, Taxes, and NotAssigned excluded. Top 10 categories shown as % of total.",
        "Category classification is set at property level in Mews. Categories include FoodAndBeverage, Accommodation, Wellness, Facilities, Events, Sport, Tourism, ExternalRevenue, SundryIncome.",
    ),
    (
        "Top Upsell Categories — Avg Value (bar chart)",
        "Average upsell value per transaction in EUR by accounting category (top 10)",
        "product.marts.mrt_daily_upsells_by_product + product.dimensions.dim_pms_properties",
        "SUM(total_upsell_gross_value_eur) / SUM(count_total_upsells) grouped by accounting_category_classification. Same filters and top 10 as volume chart.",
        "Gross value used (including taxes). Same category exclusions apply.",
    ),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Dictionary"

headers = ["Graph Title", "Metrics Shown", "Catalog / Table", "Calculations Applied", "Comments"]

# Styles
header_fill   = PatternFill("solid", fgColor="E87DC2")   # brand pink
alt_fill      = PatternFill("solid", fgColor="F9F0F6")   # light mauve
header_font   = Font(bold=True, color="FFFFFF", size=11)
body_font     = Font(size=10)
thin_border   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
wrap = Alignment(wrap_text=True, vertical="top")

# Header row
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = wrap
    cell.border    = thin_border

# Data rows
for row_idx, row in enumerate(rows, 2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill()
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = body_font
        cell.fill      = fill
        cell.alignment = wrap
        cell.border    = thin_border

# Column widths
col_widths = [35, 40, 65, 80, 80]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Auto-fit row heights (approximate)
for row in ws.iter_rows(min_row=2):
    ws.row_dimensions[row[0].row].height = 80

filename = "Mews_Dashboard_Data_Dictionary.xlsx"
wb.save(filename)
print(f"✅ Saved: {filename}")